# -*- coding:utf-8 -*-
import os
import re
import datetime

import openpyxl
import requests
from urllib.parse import urlencode

from bs4 import BeautifulSoup
from selenium import webdriver


def format_date(input_date):
    """ date examples:
    8 hours ago
    30 minutes ago
    3 days ago
    17 April 2017
    17 April = 17 April {currentYear}
    7 April = 07 April {currentYear}
    # April 9, 2021 04:06 pm ET
    :param input_date: string
    :return: int 20170417
    """
    result = -1
    try:
        now = datetime.datetime.now()
        # 非常规日期校验
        days_match = re.match(r"(\d+)\s+day.+ago", input_date)
        if days_match:
            n_days_before = now - datetime.timedelta(days=int(days_match.group(1)))
            return n_days_before.strftime("%Y%m%d")
        hours_match = re.match(r"(\d+)\s+hour.+ago", input_date)
        if hours_match:
            n_hours_before = now - datetime.timedelta(hours=int(hours_match.group(1)))
            return n_hours_before.strftime("%Y%m%d")
        minutes_match = re.match(r"(\d+)\s+minute.+ago", input_date)
        if minutes_match:
            n_minutes_before = now - datetime.timedelta(minutes=int(minutes_match.group(1)))
            return n_minutes_before.strftime("%Y%m%d")

        date_slice = re.split(r"\s+", input_date.strip())
        day = date_slice[0]
        if len(day) == 1: day = "0" + day
        # 处理脏数据
        month = MONTH_DICT.get(date_slice[1])
        if not month: month = "01"
        year = "2021"
        if len(date_slice) == 2:
            year = now.strftime("%Y")
        elif len(date_slice) == 3:
            year = date_slice[2]
        result = int(f"{year}{month}{day}")
    except:
        return result


def now_timestamp(mode="timestamp"):
    result = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    if mode == "human":
        result = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ")
    return result


def project_dir():
    # D:\projects\News-Spider
    return os.path.abspath(os.path.dirname(__file__))


def translate_with_api(text, target_language="zh-CN"):
    """
    请求谷歌翻译接口，次数过多会被限制访问，大概在20次左右
    :param text:
    :param target_language: 
    :return:
    """
    result = ""
    header = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36",
        "accept": "*/*",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "zh-CN,zh;q=0.9",
        "cookie": "wsjregion=na,us; gdprApplies=false; ab_uuid=637c98ca-f18f-4518-a47a-9518c59f567d; usr_bkt=ixi4E5ylqa; cX_P=kngcgf1o240dcpev; cX_S=kngcgf1zy97m3k1z; _sp_v1_csv=null; _sp_v1_opt=1:; _sp_v1_ss=1:H4sIAAAAAAAAAItWqo5RKimOUbLKK83J0YlRSkVil4AlqmtrlXSoqiwWACMYp9h2AAAA; _sp_v1_uid=1:798:1c2fe67d-ec2d-4dce-92aa-260e10d09019; _sp_v1_lt=1:; djvideovol=1; __gads=ID=9b4582df3b958aab:T=1618337659:S=ALNI_MY_oTx3ymoSrklLAMySQBZT_hvyUA; MicrosoftApplicationsTelemetryDeviceId=6973b8e1-34df-3a2e-cb4c-58e2fa8d98e9; MicrosoftApplicationsTelemetryFirstLaunchTime=1618337660216; cX_G=cx:pgvu3625bth816cyzb3jwhgfb:2kehjb05hoy9x; AMCVS_CB68E4BA55144CAA0A4C98A5@AdobeOrg=1; s_cc=true; _ncg_id_=b22bcf7f-11af-4eae-8bcf-7c73b63884fe; _li_dcdm_c=.wsj.com; _lc2_fpi=7880a1137012--01f3672xf29pq93a2pckahezsz; _fbp=fb.1.1618337698060.946935606; OB-USER-TOKEN=38deeaf5-c504-49c6-b6c3-29fcdec2feee; has_optimizely=true; optimizelyEndUserId=oeu1618337800998r0.08981723274843434; _ncg_g_id_=375d10e9-f063-4931-a6bd-157af07bab41; ki_r=; ResponsiveConditional_initialBreakpoint=lg; djcs_route=ace07236-6631-4caa-a2af-bc35c3dde732; djcs_secid=YzMyYmZjMjktNzJkZS00ZTgyLTk3ZGUtZTMzNjk2ZTk4ZTBm.J8opVMp2kS5CFQ1aye74_GsiHcLQlvIpJF2q-Rsqqos; TR=V2-15e14b982a648ef519cac5532c476bbe86c3fa51ee3e8e779209bd9a8ba7661c; _ntv_uid=3fc25902-f7b9-43c1-946d-ae954e9fecb3; djcs_auto=M1619298224/92y758qyl9S5yL+w/zUpcoVPq+SwXHskzdk6Avf+rkF/16poPfyMUQ+d/6a+N1hQSdtN2E3xSLLLcUIgBMmGwwXn4qpuYKnDrMnjt+tRS48VCQ7nwwt63ApQtUbURrt17u6qyifUJkabLdfABwSl/ZnpNDEA+WteEH8vU+nPgu8L/2ogvUmzULCfzbdJpva92oJqCSXnVhbTPVCRhk3o0KR+wbhXCbB/QLd2p9tXbHu4E1dlR5t+bKrQYIdHqc6KCH+hpTeJPMaadLZI+7Rn5x9t9ZLN+wZqVJjB3awX6ijrL5yzDPEpmsD80GLFeQ3qHPzb0PzhMRn6TYaL1LoOIakixHK+3BgOkkB8B69Ky++I2bem2KtQ3W90LuIlq48DpBaxwpsbr45WA8mrzwAhdw==G; djcs_session=M1619369028/r8AE2KH/yGew+uyw5V82gC2r9D6LLNpD6s0YVL3nIfChLTquAS58c4cJHExzxcyQUHK3R67k85M8eEI8N0cW/wGyzhvJgD/g2zegO7lpfHrVYCZo3lHTFFxcJ1mxYswNrgHHwBvYRww0kvjB2m71I7hNDJvWOdoNKwE7d1SzI2nb0w7tguQYp7H1/o/tw5QHNdEkNryXJovKm1nwX9AYz062flCm/AZmgsC5ZS7XomfINK908SkNGNaUZtw29nzi+giiXdOu1vbc2n19M+z1Ybd5r2TEHyEUX4+h2O+BvN3iUebi9U5G/KpEHtgGTEmcUhO80kmubdFZFAnFVcyfcjyAhjIc3SRgMCXGI9WPBLU4hy0/3ZPc27SRVnAZvP7S+UTx75s/kDGXAM96KDc+64CxMWHmG7HFcDMfSar3hkuuKwem5zvoNFgTJ5N5S6V4LgTabhiZv3bau8WBWb83jZtIaItOqiYETDAuZq1Vf+tFK6cPV043mX0hTGsAjW33G; usr_prof_v2=eyJwIjp7InBzIjowLjksInEiOjAuODF9LCJjcCI6eyJlYyI6IlN0YWJsZSIsInBjIjowLjQ1NDg0LCJwc3IiOjAuMjE5OTEsInRkIjo0MCwiYWQiOjQsInFjIjo5OSwicW8iOjk5LCJzY2VuIjp7ImNoZSI6MC4zMDU3OCwiY2huIjowLjQ4MDk4LCJjaGEiOjAuMjk5MTgsImNocCI6MC40MzE3OH19LCJpYyI6NX0=; AMCV_CB68E4BA55144CAA0A4C98A5@AdobeOrg=1585540135|MCIDTS|18743|MCMID|20248205616545994651665923179814411163|MCAID|NONE|MCOPTOUT-1619378526s|NONE|MCAAMLH-1619976126|11|MCAAMB-1619976126|j8Odv6LonN4r3an7LhD3WZrU1bUpAkFkkiY1ncBR96t2PTI|vVersion|4.4.0; DJSESSION=country=us||continent=na||region=ca||city=sanjose||latitude=37.3435||longitude=-121.8887||timezone=pst||zip=95101+95103+95106+95108-95113+95115-95136+95138-95139+95141+95148+95150-95161+95164+95170+95172-95173+95190-95194+95196; ccpaApplies=true; hok_seg=8ljs5iat1cy0,8m5oogcu3a7n; _am_sp_djcsses.1fc3=*; _scid=8e41a298-cd35-440c-b4f9-e613b3dc2ef5; ki_t=1618337814635;1619372253677;1619372283386;4;11; _sp_v1_data=2:306234:1618337657:0:41:0:41:0:0:_:-1; consentUUID=146ea439-5cc1-4b68-8ea1-145f00505ec9; utag_main=v_id:0178cc71527b000370d00be3349a03073001406b00978$_sn:6$_se:5$_ss:0$_st:1619375623302$vapi_domain:wsj.com$ses_id:1619372257684;exp-session$_pn:5;exp-session$_prevpage:;exp-1619377423313; _ncg_sp_ses.5378=*; _ncg_sp_id.5378=b22bcf7f-11af-4eae-8bcf-7c73b63884fe.1618337685.6.1619373826.1618763125.bd148ec8-8f46-4a16-9793-f43e19e763e9; _rdt_uuid=1619373826336.cd4d2f89-300c-44c7-82e2-85aaa7139d29; _tq_id.TV-63639009-1.1fc3=a06c2662c213de7e.1619373827.0.1619373827..; _sctr=1|1619366400000; s_tp=2942; s_ppv=https%3A//www.wsj.com/search%3Fquery%3Dtokyo%2520china%26isToggleOn%3Dtrue%26operator%3DAND%26sort%3Ddate-desc%26duration%3D1y%26startDate%3D2021%252F04%252F01%26endDate%3D2021%252F04%252F10%26source%3Dwsjie%252Cblog%252Cwsjsitesrch%252Cwsjpro%252Cautowire%252Capfeed,25,25,722; _am_sp_djcsid.1fc3=7fc4b535-a670-457d-95df-8430916f26b4.1618337684.6.1619374065.1618764703.b86d3273-218a-419c-98c7-d0338fe3bd99"
    }
    if text and isinstance(text, str) and text.strip() != "":
        url = "https://translate.google.com/translate_a/single?"
        params = {"client": "gtx", "sl": "auto", "tl": target_language, "dt": "t", "ie": "UTF-8", "oe": "UTF-8",
                  "q": text}
        try:
            r = requests.get(url + urlencode(params), headers=header)
            if r.status_code == 200:
                # 拼接语义分割的各部分
                for item in r.json()[0]:
                    result += item[0]
        except:
            pass
    return result


def get_webdriver():
    # 模拟浏览器登录
    options = webdriver.ChromeOptions()
    # 关闭可视化
    options.add_argument('--headless')
    # 关闭图片视频加载
    options.add_argument('blink-settings=imagesEnabled=false')
    driver = webdriver.Chrome(DRIVER_PATH, options=options)
    return driver


def translate(text, target_language="zh-CN"):
    driver = get_webdriver()
    result = ""
    for i in range(2):
        result = translate_with_webdriver(text, driver, target_language=target_language)
        if result.strip() != "":
            break
    driver.quit()
    return result


def translate_with_webdriver(text, driver, target_language="zh-CN"):
    result = ""
    if text and isinstance(text, str) and text.strip() != "":
        # 超过5000字符需要多次翻译
        if len(text) > 5000:
            count = len(text) // 5000 + 1
            temp_result = ""
            for i in range(0, count):
                temp_text = text[5000 * i:5000 * (i + 1)]
                temp_result += translate_with_webdriver(temp_text, driver, target_language)
            return temp_result
        url = f"https://translate.google.cn/?"
        params = {"sl": "auto", "tl": target_language, "op": "translate", "text": text}
        try:
            driver.get(url + urlencode(params))
            element = driver.find_element_by_xpath(
                "//div[1]/div[2]/c-wiz[1]/div[2]/c-wiz[1]/div[1]/div[2]/div[2]/c-wiz[2]/div[5]/div[1]/div[1]/span[1]/span[1]")
            soup = BeautifulSoup(element.get_attribute('innerHTML'), "html.parser")
            result = soup.find_all("span")[0].string
        except Exception as exc:
            pass
    return result


def create_xlsx_with_head(file_path, sheet_name, head_values=None):
    if head_values is None:
        head_values = ["Title", "Title-CN", "Date", "URL", "Text", "Text-CN"]
    if file_path.endswith("xlsx"):
        dir_name = os.path.dirname(file_path)
        if not os.path.isdir(dir_name):
            os.makedirs(dir_name)
        # 新建文件和表格
        wb = openpyxl.Workbook()
        # 删除默认生成的sheet
        wb.remove(wb["Sheet"])
        sheet = wb.create_sheet(sheet_name, index=1)
        # 调整列宽
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 20.0
        sheet.column_dimensions['E'].width = 50.0
        sheet.column_dimensions['F'].width = 80.0
        # 添加表头
        if head_values: sheet.append(head_values)
        wb.save(file_path)  # 保存文件，注意以xlsx为文件扩展名


def write_xlsx_apend(file_path, values):
    if os.path.isfile(file_path) and file_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file_path)
        # 获取workbook中第一个表格
        sheet = wb[wb.sheetnames[0]]
        for value in values:
            row = [value.title, value.title_cn, value.date, value.url, value.text, value.text_cn]
            sheet.append(row)
        wb.save(file_path)  # 保存文件，注意以xlsx为文件扩展名


def get_title_time_content(url, header=None):
    """
    简单解析 HTML 网页，获取新闻的文章标题、发布时间和内容，作为补充方法
    :param url:
    :return:
    """
    title = publish_date = content = ""
    try:
        r = requests.get(url, headers=header)
        if r.status_code != 200:
            return

        soup = BeautifulSoup(r.text, 'html.parser')
        # 获取标题、时间、内容
        title = soup.h1.text
        p_list = soup.find_all('p')
        content = ''
        for p in p_list:
            content += p.text + '\n'
        publish_date = soup.time.text
    except Exception as exc:
        pass
    finally:
        return title, publish_date, content


MONTH_DICT = {"January": "01",
              "February": "02",
              "March": "03",
              "April": "04",
              "May": "05",
              "June": "06",
              "July": "07",
              "August": "08",
              "September": "09",
              "October": "10",
              "November": "11",
              "December": "12"}

SUPPORT_LANGUAGE = {'afrikaans': 'af',
                    'arabic': 'ar',
                    'belarusian': 'be',
                    'bulgarian': 'bg',
                    'catalan': 'ca',
                    'czech': 'cs',
                    'welsh': 'cy',
                    'danish': 'da',
                    'german': 'de',
                    'greek': 'el',
                    'english': 'en',
                    'esperanto': 'eo',
                    'spanish': 'es',
                    'estonian': 'et',
                    'persian': 'fa',
                    'finnish': 'fi',
                    'french': 'fr',
                    'irish': 'ga',
                    'galician': 'gl',
                    'hindi': 'hi',
                    'croatian': 'hr',
                    'hungarian': 'hu',
                    'indonesian': 'id',
                    'icelandic': 'is',
                    'italian': 'it',
                    'hebrew': 'iw',
                    'japanese': 'ja',
                    'korean': 'ko',
                    'latin': 'la',
                    'lithuanian': 'lt',
                    'latvian': 'lv',
                    'macedonian': 'mk',
                    'malay': 'ms',
                    'maltese': 'mt',
                    'dutch': 'nl',
                    'norwegian': 'no',
                    'polish': 'pl',
                    'portuguese': 'pt',
                    'romanian': 'ro',
                    'russian': 'ru',
                    'slovak': 'sk',
                    'slovenian': 'sl',
                    'albanian': 'sq',
                    'serbian': 'sr',
                    'swedish': 'sv',
                    'swahili': 'sw',
                    'thai': 'th',
                    'filipino': 'tl',
                    'turkish': 'tr',
                    'ukrainian': 'uk',
                    'vietnamese': 'vi',
                    'yiddish': 'yi',
                    'chinese_simplified': 'zh-CN',
                    'chinese_traditional': 'zh-TW',
                    'auto': 'auto'}

DRIVER_PATH = os.path.join(project_dir(), "dependency/chromedriver.exe")

if __name__ == '__main__':
    # value_title = [["姓名", "性别", "年龄", "城市", "职业"], ]
    #
    # value1 = [["张三", "男", "19", "杭州", "研发工程师"],
    #           ["李四", "男", "22", "北京", "医生"],
    #           ["王五", "女", "33", "珠海", "出租车司机"], ]
    #
    # value2 = [["Tom", "男", "21", "西安", "测试工程师"],
    #           ["Jones", "女", "34", "上海", "产品经理"],
    #           ["Cat", "女", "56", "上海", "教师"], ]
    # print(format_date("17 Aprxil 2017"))
    # write_excel_xls(book_name_xls, sheet_name_xls, value_title)
    # write_excel_xls_append(book_name_xls, value1)
    # write_excel_xls_append(book_name_xls, value2)
    # read_excel_xls(book_name_xls)
    # create_xlsx_with_head(r"D:\projects\Spider\demo.xlsx", "demosheet", value_title[0])
    # write_xlsx_apend(r"D:\projects\Spider\demo.xlsx", value1)
    text = ""

    header = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36",
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "zh-CN,zh;q=0.9",
        "cache-control": "max-age=0",
        "upgrade-insecure-requests": "1",
        "Connection": "close",
        "cookie": "wsjregion=na,us; gdprApplies=false; ab_uuid=637c98ca-f18f-4518-a47a-9518c59f567d; usr_bkt=ixi4E5ylqa; cX_P=kngcgf1o240dcpev; cX_S=kngcgf1zy97m3k1z; _sp_v1_csv=null; _sp_v1_opt=1:; _sp_v1_ss=1:H4sIAAAAAAAAAItWqo5RKimOUbLKK83J0YlRSkVil4AlqmtrlXSoqiwWACMYp9h2AAAA; _sp_v1_uid=1:798:1c2fe67d-ec2d-4dce-92aa-260e10d09019; _sp_v1_lt=1:; djvideovol=1; __gads=ID=9b4582df3b958aab:T=1618337659:S=ALNI_MY_oTx3ymoSrklLAMySQBZT_hvyUA; MicrosoftApplicationsTelemetryDeviceId=6973b8e1-34df-3a2e-cb4c-58e2fa8d98e9; MicrosoftApplicationsTelemetryFirstLaunchTime=1618337660216; cX_G=cx:pgvu3625bth816cyzb3jwhgfb:2kehjb05hoy9x; AMCVS_CB68E4BA55144CAA0A4C98A5@AdobeOrg=1; s_cc=true; _ncg_id_=b22bcf7f-11af-4eae-8bcf-7c73b63884fe; _li_dcdm_c=.wsj.com; _lc2_fpi=7880a1137012--01f3672xf29pq93a2pckahezsz; _fbp=fb.1.1618337698060.946935606; OB-USER-TOKEN=38deeaf5-c504-49c6-b6c3-29fcdec2feee; has_optimizely=true; optimizelyEndUserId=oeu1618337800998r0.08981723274843434; _ncg_g_id_=375d10e9-f063-4931-a6bd-157af07bab41; ki_r=; ResponsiveConditional_initialBreakpoint=lg; djcs_route=ace07236-6631-4caa-a2af-bc35c3dde732; djcs_secid=YzMyYmZjMjktNzJkZS00ZTgyLTk3ZGUtZTMzNjk2ZTk4ZTBm.J8opVMp2kS5CFQ1aye74_GsiHcLQlvIpJF2q-Rsqqos; TR=V2-15e14b982a648ef519cac5532c476bbe86c3fa51ee3e8e779209bd9a8ba7661c; _ntv_uid=3fc25902-f7b9-43c1-946d-ae954e9fecb3; djcs_auto=M1619298224/92y758qyl9S5yL+w/zUpcoVPq+SwXHskzdk6Avf+rkF/16poPfyMUQ+d/6a+N1hQSdtN2E3xSLLLcUIgBMmGwwXn4qpuYKnDrMnjt+tRS48VCQ7nwwt63ApQtUbURrt17u6qyifUJkabLdfABwSl/ZnpNDEA+WteEH8vU+nPgu8L/2ogvUmzULCfzbdJpva92oJqCSXnVhbTPVCRhk3o0KR+wbhXCbB/QLd2p9tXbHu4E1dlR5t+bKrQYIdHqc6KCH+hpTeJPMaadLZI+7Rn5x9t9ZLN+wZqVJjB3awX6ijrL5yzDPEpmsD80GLFeQ3qHPzb0PzhMRn6TYaL1LoOIakixHK+3BgOkkB8B69Ky++I2bem2KtQ3W90LuIlq48DpBaxwpsbr45WA8mrzwAhdw==G; djcs_session=M1619369028/r8AE2KH/yGew+uyw5V82gC2r9D6LLNpD6s0YVL3nIfChLTquAS58c4cJHExzxcyQUHK3R67k85M8eEI8N0cW/wGyzhvJgD/g2zegO7lpfHrVYCZo3lHTFFxcJ1mxYswNrgHHwBvYRww0kvjB2m71I7hNDJvWOdoNKwE7d1SzI2nb0w7tguQYp7H1/o/tw5QHNdEkNryXJovKm1nwX9AYz062flCm/AZmgsC5ZS7XomfINK908SkNGNaUZtw29nzi+giiXdOu1vbc2n19M+z1Ybd5r2TEHyEUX4+h2O+BvN3iUebi9U5G/KpEHtgGTEmcUhO80kmubdFZFAnFVcyfcjyAhjIc3SRgMCXGI9WPBLU4hy0/3ZPc27SRVnAZvP7S+UTx75s/kDGXAM96KDc+64CxMWHmG7HFcDMfSar3hkuuKwem5zvoNFgTJ5N5S6V4LgTabhiZv3bau8WBWb83jZtIaItOqiYETDAuZq1Vf+tFK6cPV043mX0hTGsAjW33G; usr_prof_v2=eyJwIjp7InBzIjowLjksInEiOjAuODF9LCJjcCI6eyJlYyI6IlN0YWJsZSIsInBjIjowLjQ1NDg0LCJwc3IiOjAuMjE5OTEsInRkIjo0MCwiYWQiOjQsInFjIjo5OSwicW8iOjk5LCJzY2VuIjp7ImNoZSI6MC4zMDU3OCwiY2huIjowLjQ4MDk4LCJjaGEiOjAuMjk5MTgsImNocCI6MC40MzE3OH19LCJpYyI6NX0=; AMCV_CB68E4BA55144CAA0A4C98A5@AdobeOrg=1585540135|MCIDTS|18743|MCMID|20248205616545994651665923179814411163|MCAID|NONE|MCOPTOUT-1619378526s|NONE|MCAAMLH-1619976126|11|MCAAMB-1619976126|j8Odv6LonN4r3an7LhD3WZrU1bUpAkFkkiY1ncBR96t2PTI|vVersion|4.4.0; DJSESSION=country=us||continent=na||region=ca||city=sanjose||latitude=37.3435||longitude=-121.8887||timezone=pst||zip=95101+95103+95106+95108-95113+95115-95136+95138-95139+95141+95148+95150-95161+95164+95170+95172-95173+95190-95194+95196; ccpaApplies=true; hok_seg=8ljs5iat1cy0,8m5oogcu3a7n; _am_sp_djcsses.1fc3=*; _scid=8e41a298-cd35-440c-b4f9-e613b3dc2ef5; ki_t=1618337814635;1619372253677;1619372283386;4;11; _sp_v1_data=2:306234:1618337657:0:41:0:41:0:0:_:-1; consentUUID=146ea439-5cc1-4b68-8ea1-145f00505ec9; utag_main=v_id:0178cc71527b000370d00be3349a03073001406b00978$_sn:6$_se:5$_ss:0$_st:1619375623302$vapi_domain:wsj.com$ses_id:1619372257684;exp-session$_pn:5;exp-session$_prevpage:;exp-1619377423313; _ncg_sp_ses.5378=*; _ncg_sp_id.5378=b22bcf7f-11af-4eae-8bcf-7c73b63884fe.1618337685.6.1619373826.1618763125.bd148ec8-8f46-4a16-9793-f43e19e763e9; _rdt_uuid=1619373826336.cd4d2f89-300c-44c7-82e2-85aaa7139d29; _tq_id.TV-63639009-1.1fc3=a06c2662c213de7e.1619373827.0.1619373827..; _sctr=1|1619366400000; s_tp=2942; s_ppv=https%3A//www.wsj.com/search%3Fquery%3Dtokyo%2520china%26isToggleOn%3Dtrue%26operator%3DAND%26sort%3Ddate-desc%26duration%3D1y%26startDate%3D2021%252F04%252F01%26endDate%3D2021%252F04%252F10%26source%3Dwsjie%252Cblog%252Cwsjsitesrch%252Cwsjpro%252Cautowire%252Capfeed,25,25,722; _am_sp_djcsid.1fc3=7fc4b535-a670-457d-95df-8430916f26b4.1618337684.6.1619374065.1618764703.b86d3273-218a-419c-98c7-d0338fe3bd99"
    }

    # url = "https://www.wsj.com/articles/russia-rolls-out-covid-19-vaccine-for-animals-11622128581?page=1"
    # title, publish_date, content = get_title_time_content(url, header=header)
    # print(title)
    # print(publish_date)
    # print(content)
    # print(translate_with_webdriver(content))
    print(project_dir())
    print(now_timestamp(mode="human"))
